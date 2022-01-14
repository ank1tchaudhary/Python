import openpyxl

my_file = openpyxl.load_workbook("test.xlsx")
product_list = my_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
product_under_10_inv = {}

# print(range(product_list.max_row))


for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # calculation number of products per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier.get(supplier_name) + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # calculation of inventory value under 10
    if inventory < 11:
        product_under_10_inv[supplier_name] = inventory

    # add value for  total inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(product_under_10_inv)

my_file.save("updated_test.xlsx")




